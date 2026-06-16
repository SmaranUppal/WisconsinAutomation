"""
WCCA Court Case Scraper
=======================
Searches https://wcca.wicourts.gov/case.html for a list of competitor business names,
filters results to the last 60 days, and appends new cases (deduped by case number)
to a persistent Excel file.

Requirements:
    pip install playwright openpyxl pandas
    playwright install chromium
"""

import asyncio
import random
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from playwright.async_api import async_playwright, TimeoutError as PWTimeout

# ─────────────────────────────────────────────
# ▶  CONFIGURE THESE BEFORE RUNNING
# ─────────────────────────────────────────────
COMPETITOR_NAMES: list[str] = [
    "JG Wentworth",
    "J.G. Wentworth",
    "J G Wentworth",
    "J. G. Wentworth",
    "DRB Capital",
    "Stone Street Capital",
    "AA Ron I",
    "Abactor",
    "Abidole",
    "Adenna Med",
    "Adventura",
    "AGPI",
    "Aikman Structured Finance",
    "Annuity Transfers",
    "Apis Management",
    "Atlas Legal Funding III",
    "AXE Finance",
    "B.A.W.21",
    "B.R. Wright",
    "BHG Structured Settlements",
    "Bifco",
    "Blue Grape",
    "Catalina Structured Funding",
    "Concordis Group",
    "Conrad Factoring",
    "Cornerstone Funding",
    "Fast Annuity Settlement Transfers",
    "FL Assignments",
    "G.D.T.R.F.B.",
    "G7 Crescenta",
    "Genex Capital",
    "GJ 123",
    "Greenwood Funding",
    "Grier I",
    "Hakstol Group",
    "Hiddenview Ent",
    "JLC Capital Funding",
    "KN Direct Capital",
    "Lane Nimitz",
    "Lasko",
    "Leaf 002",
    "Legere",
    "Lottery Funding",
    "M McDougall",
    "Majestic Funding",
    "Mic-Bry8",
    "Olive Branch Funding",
    "Palermo Group",
    "Palm Green Closing",
    "Palm Harbor",
    "Passira Mal",
    "Patriot Settlement Resources",
    "QLS Funding",
    "Reliance Funding",
    "Rocorp",
    "RSL Funding",
    "Savannah Settlements",
    "Sempra Finance",
    "Seneca Originations",
    "SeneOne",
    "Settlement Capital",
    "Settlement Status",
    "Somerton",
    "Stratcap Investments",
    "Stratton Asset Funding",
    "Structured Asset Funding",
    "TKD",
    "TRM V",
    "Tybenz",
    "Uber Funding",
    "Vintage Equity Group",
    "Wepaymore Funding",
    "Zakho Way",
    "Great Plains Management",
    "T ENE",
    "RD FITZ",
    "GA OFF",
    "Assured Management",
    "Bentzen Financial",
    "Novation Funding",
    "PROM RYAN",
    "GARDEN GATE HOLDINGS",
    "AMERICAN ANNUITY FUNDING",
    "Robin Hood Funding",
    "EVERMORE ASSIGNMENTS",
    "CLOUD PEAK PARTNERS",
    "INTELIFUND",
    "UPRATE MONITOR",
    "SETTLEMENT ASSOCIATES",
    "RNSC FIN",
    "FLOWAFFIRM 9",
    "GRAYSTONE FUNDING",
    "ASSET SECURITY CONSULTING",
    "LABYRINTH FUNDING",
    "W.R. HOLDINGS",
    "AMBERVERSE",
    "HUDSON FINANCE GROUP",
    "B B BURL",
    "17-Cubs",
    "Vis-10N",
]

OUTPUT_FILE   = Path("wcca_cases.xlsx")
SHEET_NAME    = "Cases"
LOOKBACK_DAYS = 60

# Path to your real Chrome user data — so it uses your actual profile/cookies
CHROME_USER_DATA = Path.home() / "AppData" / "Local" / "Google" / "Chrome" / "User Data"

# ── Timing config (all in seconds unless noted) ──────────────────────────────
DELAY_BETWEEN_SEARCHES_MIN  = 3
DELAY_BETWEEN_SEARCHES_MAX  = 5
DELAY_AFTER_PAGE_LOAD_MIN   = .5
DELAY_AFTER_PAGE_LOAD_MAX   = 2
DELAY_AFTER_DISCLAIMER_MIN  = 1
DELAY_AFTER_DISCLAIMER_MAX  = 3
DELAY_BEFORE_TYPING_MIN     = 1
DELAY_BEFORE_TYPING_MAX     = 2
KEYSTROKE_DELAY_MIN         = 100  # ms
KEYSTROKE_DELAY_MAX         = 200  # ms
DELAY_AFTER_ENTER_MIN       = 1
DELAY_AFTER_ENTER_MAX       = 3
DELAY_AFTER_RESULTS_MIN     = 1
DELAY_AFTER_RESULTS_MAX     = 3
DELAY_CAPTCHA_RESTART       = 5
LONG_BREAK_EVERY_N          = 4
LONG_BREAK_MIN              = 5
LONG_BREAK_MAX              = 10
# ─────────────────────────────────────────────────────────────────────────────

WCCA_URL = "https://wcca.wicourts.gov/case.html"
CUTOFF   = datetime.today() - timedelta(days=LOOKBACK_DAYS)
COLUMNS  = ["Case Number", "Filing Date", "County", "Name", "Caption", "Search Term"]
DATE_FMT = "%m-%d-%Y"


def _rnd(lo, hi):
    return random.uniform(lo, hi)

def _rnd_ms(lo, hi):
    return random.randint(lo, hi)


# ── helpers ───────────────────────────────────

def _init_workbook() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    col_widths   = [18, 14, 18, 30, 50, 22]
    for col_idx, (col_name, width) in enumerate(zip(COLUMNS, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"
    wb.save(OUTPUT_FILE)
    print(f"[init] Created {OUTPUT_FILE}")


def _load_existing_case_numbers() -> set[str]:
    if not OUTPUT_FILE.exists():
        return set()
    df = pd.read_excel(OUTPUT_FILE, sheet_name=SHEET_NAME, dtype=str)
    if "Case Number" not in df.columns:
        return set()
    return set(df["Case Number"].dropna().str.strip())


def _append_rows(rows: list[dict]) -> int:
    if not rows:
        return 0
    wb = load_workbook(OUTPUT_FILE)
    ws = wb[SHEET_NAME]
    next_row = ws.max_row + 1
    row_font = Font(name="Arial", size=10)
    alt_fill = PatternFill("solid", fgColor="DCE6F1")
    written  = 0
    for r in rows:
        fill = alt_fill if (next_row % 2 == 0) else None
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=r.get(col_name, ""))
            cell.font      = row_font
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
        ws.row_dimensions[next_row].height = 15
        next_row += 1
        written  += 1
    wb.save(OUTPUT_FILE)
    return written


def _parse_date(date_str: str) -> datetime | None:
    try:
        return datetime.strptime(date_str.strip(), DATE_FMT)
    except ValueError:
        return None


async def _is_captcha(page) -> bool:
    title = (await page.title()).lower()
    title_triggers = ["captcha", "access denied", "blocked", "verify you are human", "unusual traffic"]
    if any(t in title for t in title_triggers):
        return True
    has_input       = await page.locator('input[name="businessName"]').count() > 0
    has_table       = await page.locator("#caseSearchResults").count() > 0
    has_detail_page = await page.locator(".caseNo").count() > 0
    # Valid states: search form, results table, or single-case detail page
    if not has_input and not has_table and not has_detail_page:
        print(f"  [debug] Unexpected page — title: '{await page.title()}'")
        return True
    return False


async def _new_context_and_page(playwright):
    """Launch a real Chrome instance and go straight to WCCA."""
    context = await playwright.chromium.launch_persistent_context(
        user_data_dir=str(Path.home() / "AppData" / "Local" / "WisconsinScraper" / "Profile"),
        channel="chrome",
        headless=False,
        no_viewport=True,
        args=["--start-maximized", "--disable-blink-features=AutomationControlled"],
        ignore_default_args=["--enable-automation", "--no-sandbox"],
    )

    # Grab the tab that launched, or make one
    page = context.pages[0] if context.pages else await context.new_page()

    # Navigate immediately so it never sits on about:blank
    await page.goto(WCCA_URL, wait_until="domcontentloaded", timeout=30_000)

    return context, page


async def _parse_detail_page(page, business_name: str) -> list[dict]:
    """Parse a single-case detail page when WCCA skips straight to it."""
    results = []

    try:
        case_num    = await page.locator(".caseNo").first.inner_text()
        caption     = await page.locator(".caption").first.inner_text()
        county      = await page.locator(".countyName").first.inner_text()
        filing_date = await page.locator(".cell-3.s-cell-12.field").first.inner_text()
    except Exception as e:
        print(f"  [!] Could not parse detail page: {e}")
        return results

    case_num    = case_num.strip()
    caption     = caption.strip()
    county      = county.strip()
    filing_date = filing_date.strip()

    print(f"  [detail page] case:{case_num} date:{filing_date} county:{county}")

    parsed = _parse_date(filing_date)
    if parsed is None or parsed < CUTOFF:
        print(f"  [info] Detail page case outside 60-day window: {filing_date}")
        return results

    results.append({
        "Case Number": case_num,
        "Filing Date": filing_date,
        "County":      county,
        "Name":        "",
        "Caption":     caption,
        "Search Term": business_name,
    })
    return results

# ── scraper ───────────────────────────────────

async def search_business(page, business_name: str) -> tuple[list[dict], bool]:
    results: list[dict] = []

    # ── Load page (skip if already on WCCA) ───
    if WCCA_URL not in page.url:
        try:
            await page.goto(WCCA_URL, wait_until="domcontentloaded", timeout=30_000)
        except PWTimeout:
            print(f"  [!] Timeout loading page for '{business_name}'")
            return results, False

    settle = _rnd(DELAY_AFTER_PAGE_LOAD_MIN, DELAY_AFTER_PAGE_LOAD_MAX)
    print(f"  [wait] {settle:.1f}s after page load…")
    await page.wait_for_timeout(int(settle * 1000))

    if await _is_captcha(page):
        print(f"  [!] CAPTCHA detected on page load for '{business_name}'")
        return results, True

    # ── Accept disclaimer if present ──────────
    try:
        disclaimer_btn = page.locator("button:has-text('I Agree'), input[value='I Agree']")
        if await disclaimer_btn.count() > 0:
            await disclaimer_btn.first.click()
            d = _rnd(DELAY_AFTER_DISCLAIMER_MIN, DELAY_AFTER_DISCLAIMER_MAX)
            print(f"  [wait] {d:.1f}s after disclaimer…")
            await page.wait_for_timeout(int(d * 1000))
    except Exception:
        pass

    # ── Pause before touching the input ───────
    pre_type = _rnd(DELAY_BEFORE_TYPING_MIN, DELAY_BEFORE_TYPING_MAX)
    print(f"  [wait] {pre_type:.1f}s before typing…")
    await page.wait_for_timeout(int(pre_type * 1000))

    # ── Fill business name field ───────────────
    try:
        biz_input = page.locator('input[name="businessName"]')
        await biz_input.wait_for(state="visible", timeout=15_000)
        await biz_input.click()
        await page.wait_for_timeout(_rnd_ms(300, 700))
        await biz_input.fill("")
        await page.wait_for_timeout(_rnd_ms(200, 500))
        await biz_input.type(business_name, delay=_rnd_ms(KEYSTROKE_DELAY_MIN, KEYSTROKE_DELAY_MAX))
    except PWTimeout:
        print(f"  [!] Could not find businessName input for '{business_name}'")
        return results, False

    # ── Pause before pressing Enter ────────────
    pre_enter = _rnd(DELAY_AFTER_ENTER_MIN, DELAY_AFTER_ENTER_MAX)
    print(f"  [wait] {pre_enter:.1f}s before pressing Enter…")
    await page.wait_for_timeout(int(pre_enter * 1000))

    await biz_input.press("Enter")

    # ── Wait for results or detail page ───────
    for attempt in range(5):
        try:
            # Wait for either the results table OR a direct detail page
            await page.wait_for_selector(
                "#caseSearchResults tbody tr, .caseNo",
                timeout=20_000
            )
            settle2 = _rnd(DELAY_AFTER_RESULTS_MIN, DELAY_AFTER_RESULTS_MAX)
            print(f"  [wait] {settle2:.1f}s after results appear…")
            await page.wait_for_timeout(int(settle2 * 1000))
            break
        except PWTimeout:
            if await _is_captcha(page):
                print(f"  [!] CAPTCHA detected while waiting for results (attempt {attempt + 1})")
                return results, True
            print(f"  [!] Attempt {attempt + 1} timed out, retrying '{business_name}'...")
            await page.wait_for_timeout(_rnd_ms(3000, 6000))
            await biz_input.fill("")
            await page.wait_for_timeout(_rnd_ms(500, 1000))
            await biz_input.type(business_name, delay=_rnd_ms(KEYSTROKE_DELAY_MIN, KEYSTROKE_DELAY_MAX))
            await page.wait_for_timeout(_rnd_ms(1000, 2000))
            await biz_input.press("Enter")
    else:
        print(f"  [!] All retries failed for '{business_name}', skipping.")
        return results, False

    if await _is_captcha(page):
        print(f"  [!] CAPTCHA detected after results loaded for '{business_name}'")
        return results, True

    # ── Check if we landed on a detail page instead of the results table ──
    if await page.locator(".caseNo").count() > 0 and await page.locator("#caseSearchResults").count() == 0:
        print(f"  [info] Single result — landed on detail page directly")
        return await _parse_detail_page(page, business_name), False

    # ── Parse result table ────────────────────
    header_cells = await page.query_selector_all("#caseSearchResults thead th")
    headers = [(await c.inner_text()).strip().lower() for c in header_cells]
    print(f"  [debug] Headers: {headers}")

    def col(fragment: str) -> int | None:
        for i, h in enumerate(headers):
            if fragment in h:
                return i
        return None

    idx_case    = col("case number") if col("case number") is not None else col("case")
    idx_date    = col("filing date") if col("filing date") is not None else col("filing")
    idx_county  = col("county")
    idx_name    = col("name")
    idx_caption = col("caption")

    print(f"  [debug] Col indices — case:{idx_case} date:{idx_date} county:{idx_county} name:{idx_name} caption:{idx_caption}")

    rows = await page.query_selector_all("#caseSearchResults tbody tr")
    print(f"  [debug] {len(rows)} data row(s) found")

    if len(rows) <= 1:
        print(f"  [info] No cases found for '{business_name}'")
        return results, False

    for row in rows:
        cells = await row.query_selector_all("td")

        async def get(idx):
            if idx is None or idx >= len(cells):
                return ""
            return (await cells[idx].inner_text()).strip()

        case_num    = await get(idx_case)
        filing_date = await get(idx_date)
        county      = await get(idx_county)
        name        = await get(idx_name)
        caption     = await get(idx_caption)

        if not case_num:
            continue

        parsed = _parse_date(filing_date)
        if parsed is None or parsed < CUTOFF:
            continue

        results.append({
            "Case Number": case_num,
            "Filing Date": filing_date,
            "County":      county,
            "Name":        name,
            "Caption":     caption,
            "Search Term": business_name,
        })

    return results, False


async def run_scraper() -> None:
    if not OUTPUT_FILE.exists():
        _init_workbook()

    existing_cases = _load_existing_case_numbers()
    print(f"[info] {len(existing_cases)} case(s) already in {OUTPUT_FILE}")
    print(f"[info] Searching {len(COMPETITOR_NAMES)} name(s) | cutoff: {CUTOFF.strftime(DATE_FMT)}")
    print(f"[info] Using real Chrome profile from: {CHROME_USER_DATA}\n")

    all_new_rows: list[dict] = []
    seen_this_run: set[str]  = set()

    async with async_playwright() as p:
        context, page = await _new_context_and_page(p)
        i = 0

        while i < len(COMPETITOR_NAMES):
            name = COMPETITOR_NAMES[i]
            print(f"[search] ({i + 1}/{len(COMPETITOR_NAMES)}) {name!r} …")

            rows, captcha_hit = await search_business(page, name)

            if captcha_hit:
                if all_new_rows:
                    saved = _append_rows(all_new_rows)
                    existing_cases.update(r["Case Number"] for r in all_new_rows)
                    all_new_rows.clear()
                    print(f"  [captcha] Saved {saved} row(s) to file before restarting.")
                print(f"  [captcha] Closing browser and waiting {DELAY_CAPTCHA_RESTART}s before resuming at '{name}'...")
                try:
                    await context.close()
                except Exception:
                    pass
                await asyncio.sleep(DELAY_CAPTCHA_RESTART)
                context, page = await _new_context_and_page(p)
                continue  # retry same index

            print(f"         found {len(rows)} case(s) in last {LOOKBACK_DAYS} days")

            for row in rows:
                case_num = row["Case Number"].strip()
                if case_num in existing_cases:
                    print(f"         skip (already in file): {case_num}")
                    continue
                if case_num in seen_this_run:
                    print(f"         skip (duplicate this run): {case_num}")
                    continue
                seen_this_run.add(case_num)
                all_new_rows.append(row)

            i += 1

            if i < len(COMPETITOR_NAMES):
                
                delay = _rnd(DELAY_BETWEEN_SEARCHES_MIN, DELAY_BETWEEN_SEARCHES_MAX)
                print(f"  [wait] {delay:.1f}s before next search…")
                await asyncio.sleep(delay)

        await context.close()

    written = _append_rows(all_new_rows)
    print(f"\n[done] Appended {written} new case(s) to {OUTPUT_FILE}")


if __name__ == "__main__":
    asyncio.run(run_scraper())